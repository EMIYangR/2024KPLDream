import json
from datetime import datetime
from openpyxl import load_workbook


def setJsonToExcel(files):
	# 打开传入的data.json
    with open(files, 'r') as file:
        data = json.load(file)
		
	# 加载Json文件的选手、教练票数和时间戳
    membervotes_data = data['jData']['membervotes']
    coachvotes_data = data['jData']['coachvotes']
    timestamp = data['jData']['nowtime']+28800
	
	# 打开工作簿
    workbook = load_workbook('2024KPL梦之队.xlsx')
	
	# 指定写入的工作表名称
    sheet1 = workbook['MemberData']
    sheet2 = workbook['CoachData']
    sheet3 = workbook['TimesTamp']
	
	# 写入Excel文件
    for row, (key, value) in enumerate(membervotes_data.items(), start=1):
        sheet1.cell(row=row, column=8, value=key)
        sheet1.cell(row=row, column=9, value=value)
    for row, (key, value) in enumerate(coachvotes_data.items(), start=1):
        sheet2.cell(row=row, column=6, value=key)
        sheet2.cell(row=row, column=7, value=value)
    sheet3.cell(row=2, column=2).value = datetime.utcfromtimestamp(timestamp).strftime('%Y-%m-%d %H:%M:%S')
	# 保存写入的值
    workbook.save('2024KPL梦之队.xlsx')
    return 0
