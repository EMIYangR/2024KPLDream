import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from openpyxl import load_workbook

# 设置 Chrome 选项
chrome_options = Options()
chrome_options.add_argument("--headless")  # 无头模式，不打开浏览器窗口
chrome_options.add_argument("--disable-gpu")  # 禁用 GPU 加速

# 启动 Chrome
driver_path = r"D:\EMI\Tools\chromedriver\chromedriver.exe"  # 根据你的环境设置 chromedriver 的路径
driver = webdriver.Chrome(executable_path=driver_path, options=chrome_options)

# 打开指定网页
driver.get("https://pvp.qq.com/cp/a20240424vote/index.html")
print("页面正在加载......")
# 等待 JavaScript 加载完成
wait = WebDriverWait(driver, 20)
time.sleep(1)
print("页面加载完成！")
# 执行 JavaScript 代码
result1 = driver.execute_script('''
    let all = "";
    $.each(teams, function (index, data) {
        let tname = data.tname;
        $.each(data.members, function (index, data) {
            all = all + data.nick + "\t" + data.road + "\t" + data.vote + "\t" + tname + "\\n";
        });
    });
    return all;
''')

result2 = driver.execute_script('''
    let all = "";
    $.each(coachs, function (index, data) {
        all = all + data.name + "\t" + data.vote + "\\n";
    });
    return all;
''')

# 输出结果
print("调取数据成功！")

# 关闭浏览器
driver.quit()

# 打开已有的Excel文件
workbook = load_workbook('2024KPL梦之队.xlsx')

# 获取指定的工作表
sheet1 = workbook['MemberData']  # 选手
sheet2 = workbook['CoachData']  # 教练
sheet3 = workbook['timestamp']  # 时间戳

# 将数据分割成行，并逐行写入指定的单元格
for i, line in enumerate(result1.split('\n')):
    for j, value in enumerate(line.split('\t')):
        sheet1.cell(row=i + 1, column=j + 1).value = value

for i, line in enumerate(result2.split('\n')):
    for j, value in enumerate(line.split('\t')):
        sheet2.cell(row=i + 1, column=j + 1).value = value

# 获取当前时间
timestamp = datetime.fromtimestamp(datetime.now().timestamp()).strftime('%Y-%m-%d %H:%M:%S')
sheet3.cell(row=1, column=1).value = timestamp

# 保存修改后的工作簿到文件
workbook.save('2024KPL梦之队.xlsx')

print("写入数据成功！")
